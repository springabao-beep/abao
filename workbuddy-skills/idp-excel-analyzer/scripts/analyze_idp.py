import os
import sys
import argparse

# Parse arguments before importing matplotlib so MPLCONFIGDIR can be set based on output_dir
parser = argparse.ArgumentParser(description='Analyze IDP Excel file')
parser.add_argument('input_file', help='Path to IDP Excel file')
parser.add_argument('output_dir', nargs='?', help='Output directory (default: same dir as input / Final)')
args = parser.parse_args()

INPUT_FILE = args.input_file
if args.output_dir:
    OUTPUT_DIR = args.output_dir
else:
    # Default to <input_parent>/Final
    OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(INPUT_FILE)), 'Final')

# Set matplotlib config dir before importing matplotlib to avoid AppData access
mpl_config_dir = os.path.join(OUTPUT_DIR, '.mpl_config')
os.environ['MPLCONFIGDIR'] = mpl_config_dir
os.makedirs(os.environ['MPLCONFIGDIR'], exist_ok=True)

CHART_DIR = os.path.join(OUTPUT_DIR, "charts")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(CHART_DIR, exist_ok=True)

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib import font_manager as fm_manager
import jieba
import jieba.analyse
import re
from collections import Counter
from datetime import datetime, timedelta

# Configure stdout
sys.stdout.reconfigure(encoding='utf-8')

# Matplotlib Chinese font setup
chinese_font = None
font_candidates = ['Microsoft YaHei', 'SimHei', 'SimSun', 'DengXian', 'PingFang SC']
for candidate in font_candidates:
    try:
        font_path = fm_manager.findfont(candidate, fallback_to_default=False)
        if font_path and candidate.lower() in font_path.lower() or candidate.replace(' ', '').lower() in font_path.lower():
            chinese_font = candidate
            break
    except Exception:
        continue

if chinese_font is None:
    # Try to find any font file with CJK support by checking file names
    system_fonts = fm_manager.findSystemFonts(fontpaths=None, fontext='ttf')
    for fp in system_fonts:
        if any(k in fp.lower() for k in ['yahei', 'simhei', 'simsun', 'dengxian', 'pingfang', 'noto']):
            chinese_font = fm.FontProperties(fname=fp).get_name()
            break

if chinese_font:
    plt.rcParams['font.sans-serif'] = [chinese_font]
    plt.rcParams['axes.unicode_minus'] = False
    print(f"Using font: {chinese_font}")
else:
    print("Warning: No Chinese font found. Charts may show garbled text.")

# ============================================================
# 2. Read & Clean Data
# ============================================================
print("Reading Excel file...")
df = pd.read_excel(INPUT_FILE, sheet_name=0, header=2)

# Drop rows where 工号 is missing (total row or empty rows)
df = df[df['工号'].notna()].copy()

# Standardize column names
column_map = {
    '工号': '工号',
    '员工姓名': '员工姓名',
    '经理姓名': '经理姓名',
    '目前职位': '目前职位',
    '店铺': '店铺',
    '日期': '填写日期',
    '请勾选发展目标1': '目标1类型',
    '目标1的详细描述': '目标1描述',
    '目标1的行动计划（70%-工作中实践）': '目标1_70计划',
    '目标1的行动描述（70%-工作中实践）': '目标1_70描述',
    '目标1的行动计划（20%-向他人学习）': '目标1_20计划',
    '目标1的行动描述（20%-向他人学习）': '目标1_20描述',
    '目标1的行动计划（10%-正式课程）': '目标1_10计划',
    '目标1的行动描述（10%-正式课程）': '目标1_10描述',
    '目标1的其它所需资源': '目标1资源',
    '请勾选发展目标2': '目标2类型',
    '目标2的详细描述': '目标2描述',
    '目标2的行动计划（70%-工作中实践）': '目标2_70计划',
    '目标2的行动描述（70%-工作中实践）': '目标2_70描述',
    '目标2的行动计划（20%-向他人学习）': '目标2_20计划',
    '目标2的行动描述（20%-向他人学习）': '目标2_20描述',
    '目标2的行动计划（10%-正式课程）': '目标2_10计划',
    '目标2的行动描述（10%-正式课程）': '目标2_10描述',
    '目标2的其它需要资源': '目标2资源',
    '您是否接受跨城市/跨部门的调动？': '接受调动',
    '请写明您想调动至哪里？': '调动目标',
    '总目标达成时间': '总目标达成时间',
    '中期回顾时间': '中期回顾时间',
    '终期回顾时间': '终期回顾时间'
}

df.rename(columns=column_map, inplace=True)

# Trim whitespace on text columns
for col in ['员工姓名', '经理姓名', '目前职位', '店铺', '目标1类型', '目标2类型', '接受调动']:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace('nan', np.nan)

# Normalize position names
def normalize_position(pos):
    if pd.isna(pos):
        return pos
    p = str(pos).strip().lower().replace(' ', '')
    if p in ['ssa', 'seniorsalesassociate']:
        return 'SSA'
    elif p in ['sa', 'sales', 'seniorsales', 'salesassociate']:
        return 'SA'
    elif p in ['supervisor', 'supvisor', 'spv', '主管', 'superviser']:
        return 'Supervisor'
    elif p in ['admin', 'administrator']:
        return 'Admin'
    elif p in ['senioradmin', 'senioradministrator']:
        return 'Senior Admin'
    elif p in ['eliteclientsupervisor']:
        return 'Elite Client Supervisor'
    else:
        return str(pos).strip()

df['目前职位'] = df['目前职位'].apply(normalize_position)

# Parse dates
def parse_chinese_date(x):
    if pd.isna(x) or x == 'nan':
        return pd.NaT
    if isinstance(x, pd.Timestamp):
        return x
    try:
        # Remove Chinese characters and parse
        s = str(x).strip()
        s = s.replace('年', '-').replace('月', '-').replace('日', '').strip('-')
        return pd.to_datetime(s, errors='coerce')
    except:
        return pd.NaT

for col in ['填写日期', '总目标达成时间', '中期回顾时间', '终期回顾时间']:
    df[col] = df[col].apply(parse_chinese_date)

# Fill missing goal types with NaN
df['目标1类型'] = df['目标1类型'].replace('', np.nan)
df['目标2类型'] = df['目标2类型'].replace('', np.nan)

print(f"Loaded {len(df)} employee records.")

# ============================================================
# 3. Derived Fields
# ============================================================
# Goal existence flags
df['目标1已填'] = df['目标1类型'].notna()
df['目标2已填'] = df['目标2类型'].notna()
df['目标总数'] = df['目标1已填'].astype(int) + df['目标2已填'].astype(int)

# Skill type classification (soft vs hard) from parentheses
def extract_skill_type(goal_type):
    if pd.isna(goal_type):
        return '未填写'
    if '软性' in goal_type:
        return '软性技能'
    elif '硬性' in goal_type:
        return '硬性技能'
    else:
        return '其他'

df['目标1技能属性'] = df['目标1类型'].apply(extract_skill_type)
df['目标2技能属性'] = df['目标2类型'].apply(extract_skill_type)

# Goal theme classification
def classify_goal_theme(goal_type):
    if pd.isna(goal_type):
        return '未填写'
    gt = goal_type
    if '沟通' in gt:
        return '沟通'
    elif '计划' in gt or '组织' in gt:
        return '计划与组织'
    elif '销售' in gt:
        return '销售艺术'
    elif '客户' in gt or '卓越' in gt:
        return '客户体验'
    elif '品牌' in gt or '文化' in gt:
        return '品牌与文化'
    elif '结果' in gt or '结果导向' in gt:
        return '结果导向'
    elif '领导' in gt or '团队' in gt:
        return '领导力/团队'
    elif '数据分析' in gt or '数据' in gt:
        return '数据分析'
    elif '管理' in gt:
        return '管理'
    else:
        return '其他'

df['目标1主题'] = df['目标1类型'].apply(classify_goal_theme)
df['目标2主题'] = df['目标2类型'].apply(classify_goal_theme)

# Action plan coverage (70/20/10) per goal
def action_filled(plan, desc):
    return (pd.notna(plan) and str(plan).strip() != '') or (pd.notna(desc) and str(desc).strip() != '')

for goal in ['目标1', '目标2']:
    for pct in ['70', '20', '10']:
        plan_col = f'{goal}_{pct}计划'
        desc_col = f'{goal}_{pct}描述'
        df[f'{goal}_{pct}已填'] = df.apply(lambda row: action_filled(row.get(plan_col), row.get(desc_col)), axis=1)

# Action methods extracted from plan columns
action_method_keywords = {
    '协助新任务/项目': ['协助新任务', '协助任务'],
    '负责新任务/项目': ['负责新任务', '负责项目'],
    '协助解决陷入困境的任务': ['协助解决', '陷入困境'],
    '主持店铺会议/展示成果': ['主持店铺会议', '展示团队', '展示成果'],
    '参加跨小组/团队/店铺/职能会议': ['参加跨小组', '跨团队', '跨店铺', '职能会议'],
    '评估并优化现有流程': ['评估并优化', '优化现有流程'],
    '观察优秀同事': ['观察优秀同事', '观察优秀'],
    '辅导': ['辅导'],
    '向经理寻求反馈': ['向经理寻求反馈', '经理反馈'],
    '人脉交流学习': ['人脉交流', '交流学习'],
    '线上学习（哈佛管理导师平台）': ['线上学习', '哈佛管理'],
    '线下培训（集团/公司培训、工作坊）': ['线下培训', '公司培训', '工作坊'],
    '自主学习（阅读书籍、收听播客、学习语言等）': ['自主学习', '阅读书籍', '收听播客', '学习语言']
}

def extract_action_methods(text):
    if pd.isna(text):
        return []
    text_str = str(text)
    found = []
    for method, keywords in action_method_keywords.items():
        if any(kw in text_str for kw in keywords):
            found.append(method)
    return found

# Extract action methods for all plan columns
all_action_methods = []
for goal in ['目标1', '目标2']:
    for pct in ['70', '20', '10']:
        col = f'{goal}_{pct}计划'
        df[f'{goal}_{pct}方法'] = df[col].apply(extract_action_methods)
        all_action_methods.extend(df[f'{goal}_{pct}方法'].tolist())

# Flatten action methods
action_method_list = []
for methods in all_action_methods:
    action_method_list.extend(methods)
action_method_counts = Counter(action_method_list)

# Resource columns filled
df['目标1资源已填'] = df['目标1资源'].apply(lambda x: pd.notna(x) and str(x).strip() != '')
df['目标2资源已填'] = df['目标2资源'].apply(lambda x: pd.notna(x) and str(x).strip() != '')

# Completeness score
# Score: 1 point per goal type, 1 per goal description, 1 per action dimension (plan or desc), 1 per resource
# Max per goal = 1 + 1 + 3 + 1 = 6
# Max per employee = 12

df['目标1描述已填'] = df['目标1描述'].apply(lambda x: pd.notna(x) and str(x).strip() != '')
df['目标2描述已填'] = df['目标2描述'].apply(lambda x: pd.notna(x) and str(x).strip() != '')

df['完整度得分'] = (
    df['目标1已填'].astype(int) +
    df['目标1描述已填'].astype(int) +
    df['目标1_70已填'].astype(int) +
    df['目标1_20已填'].astype(int) +
    df['目标1_10已填'].astype(int) +
    df['目标1资源已填'].astype(int) +
    df['目标2已填'].astype(int) +
    df['目标2描述已填'].astype(int) +
    df['目标2_70已填'].astype(int) +
    df['目标2_20已填'].astype(int) +
    df['目标2_10已填'].astype(int) +
    df['目标2资源已填'].astype(int)
)

df['完整度百分比'] = (df['完整度得分'] / 12 * 100).round(1)

def completeness_tier(score):
    if score >= 10:
        return '高'
    elif score >= 6:
        return '中'
    else:
        return '低'

df['完整度等级'] = df['完整度得分'].apply(completeness_tier)

# Cycle analysis
df['填写到总目标天数'] = (df['总目标达成时间'] - df['填写日期']).dt.days
df['中期到终期天数'] = (df['终期回顾时间'] - df['中期回顾时间']).dt.days
df['终期到总目标天数'] = (df['总目标达成时间'] - df['终期回顾时间']).dt.days

# Anomaly flags
df['异常_总目标早于填写'] = df['填写到总目标天数'] < 0
df['异常_中期晚于终期'] = df['中期到终期天数'] < 0
df['异常_终期晚于总目标'] = df['终期到总目标天数'] < 0

# ============================================================
# 4. Analysis Summaries
# ============================================================
print("Computing summaries...")

# --- Basic situation ---
num_employees = len(df)
num_managers = df['经理姓名'].nunique()
num_stores = df['店铺'].nunique()
num_positions = df['目前职位'].nunique()

position_dist = df['目前职位'].value_counts().reset_index()
position_dist.columns = ['目前职位', '人数']

manager_dist = df['经理姓名'].value_counts().reset_index()
manager_dist.columns = ['经理姓名', '下属人数']

store_dist = df['店铺'].value_counts().reset_index()
store_dist.columns = ['店铺', '人数']

mobility_dist = df['接受调动'].value_counts().reset_index()
mobility_dist.columns = ['接受调动', '人数']

fill_date_dist = df['填写日期'].dt.to_period('W').value_counts().sort_index().reset_index()
fill_date_dist.columns = ['周', '人数']
fill_date_dist['周'] = fill_date_dist['周'].astype(str)

# --- IDP goal analysis ---
goal1_types = df['目标1类型'].value_counts(dropna=False).reset_index()
goal1_types.columns = ['发展目标类型', '人数']
goal1_types['发展目标类型'] = goal1_types['发展目标类型'].fillna('未填写')

goal2_types = df['目标2类型'].value_counts(dropna=False).reset_index()
goal2_types.columns = ['发展目标类型', '人数']
goal2_types['发展目标类型'] = goal2_types['发展目标类型'].fillna('未填写')

# Combine goal types across both goals
all_goals = pd.concat([
    df['目标1类型'].rename('发展目标类型'),
    df['目标2类型'].rename('发展目标类型')
])
all_goals = all_goals[all_goals.notna()]
goal_type_combined = all_goals.value_counts().reset_index()
goal_type_combined.columns = ['发展目标类型', '出现次数']

# Skill attribute distribution
skill_attr_dist = pd.concat([
    df['目标1技能属性'].rename('技能属性'),
    df['目标2技能属性'].rename('技能属性')
]).value_counts().reset_index()
skill_attr_dist.columns = ['技能属性', '出现次数']

# Theme distribution
theme_dist = pd.concat([
    df['目标1主题'].rename('主题'),
    df['目标2主题'].rename('主题')
]).value_counts().reset_index()
theme_dist.columns = ['主题', '出现次数']

# Goal completeness categories
def goal_completeness_category(row):
    g1 = row['目标1已填']
    g2 = row['目标2已填']
    if g1 and g2:
        return '两个目标均填写'
    elif g1 and not g2:
        return '仅填写目标1'
    elif not g1 and g2:
        return '仅填写目标2'
    else:
        return '两个目标均未填写'

df['目标填写情况'] = df.apply(goal_completeness_category, axis=1)
goal_completeness_dist = df['目标填写情况'].value_counts().reset_index()
goal_completeness_dist.columns = ['目标填写情况', '人数']

# Keyword extraction using jieba
stopwords = set(['的', '了', '在', '是', '和', '与', '及', '等', '对', '为', '有', '我', '要', '将', '并', '以', '以及', '通过', '提升', '学习', '能力', '工作', '目标', '计划', '行动', '提高', '加强', '优化', '进行', '需要', '可以', '能够', '希望', '做到', '达到', '完成', '实现', '帮助', '协助', '负责', '每周', '每日', '每月', '每季度', '一次', '一个', '一些', '不同', '相关', '有关', '对于', '关于', '根据', '按照', '结合', '针对', '不断', '持续', '进一步', '更好', '更多', '更加', '有效', '高效', '良好', '积极', '主动', '及时', '定期', '充分', '全面', '深入', '系统', '专业', '实际', '具体', '明确', '清晰', '合理', '科学', '规范', '标准', '流程', '方法', '方式', '措施', '方案', '路径', '渠道', '平台', '资源', '支持', '反馈', '指导', '沟通', '交流', '分享', '汇报', '总结', '复盘', '反思', '改进', '改善', '完善', '落实', '执行', '推进', '推动', '促进', '带动', '确保', '保障', '维护', '管理', '经营', '运营', '服务', '销售', '客户', '团队', '员工', '同事', '经理', '领导', '店铺', '公司', '集团', '部门', '区域', '门店', '柜台', '业务', '项目', '任务', '活动', '会议', '课程', '培训', '书籍', '播客', '语言'])

# Extract keywords from goal descriptions
desc_text = ' '.join(df['目标1描述'].fillna('').astype(str)) + ' ' + ' '.join(df['目标2描述'].fillna('').astype(str))
words = jieba.lcut(desc_text)
words = [w for w in words if len(w) >= 2 and w not in stopwords and not re.match(r'[a-zA-Z0-9]+', w)]
keyword_counts = Counter(words).most_common(30)
keyword_df = pd.DataFrame(keyword_counts, columns=['关键词', '出现频次'])

# --- Action plan analysis ---
action_coverage = pd.DataFrame({
    '维度': ['70% 工作中实践', '20% 向他人学习', '10% 正式课程'],
    '目标1填写数': [
        df['目标1_70已填'].sum(),
        df['目标1_20已填'].sum(),
        df['目标1_10已填'].sum()
    ],
    '目标1填写率': [
        f"{df['目标1_70已填'].sum() / len(df) * 100:.1f}%",
        f"{df['目标1_20已填'].sum() / len(df) * 100:.1f}%",
        f"{df['目标1_10已填'].sum() / len(df) * 100:.1f}%"
    ],
    '目标2填写数': [
        df['目标2_70已填'].sum(),
        df['目标2_20已填'].sum(),
        df['目标2_10已填'].sum()
    ],
    '目标2填写率': [
        f"{df['目标2_70已填'].sum() / len(df) * 100:.1f}%",
        f"{df['目标2_20已填'].sum() / len(df) * 100:.1f}%",
        f"{df['目标2_10已填'].sum() / len(df) * 100:.1f}%"
    ]
})

action_method_df = pd.DataFrame(action_method_counts.most_common(), columns=['行动方式', '出现次数'])

resource_filled_rate = ((df['目标1资源已填'].sum() + df['目标2资源已填'].sum()) / (len(df) * 2) * 100)
resource_text = ' '.join(df['目标1资源'].fillna('').astype(str)) + ' ' + ' '.join(df['目标2资源'].fillna('').astype(str))
resource_words = jieba.lcut(resource_text)
resource_words = [w for w in resource_words if len(w) >= 2 and w not in stopwords and not re.match(r'[a-zA-Z0-9]+', w)]
resource_keyword_counts = Counter(resource_words).most_common(20)
resource_keyword_df = pd.DataFrame(resource_keyword_counts, columns=['资源关键词', '出现频次'])

completeness_dist = df['完整度等级'].value_counts().reset_index()
completeness_dist.columns = ['完整度等级', '人数']

manager_supported_methods = ['向经理寻求反馈', '辅导', '人脉交流学习']
manager_support_count = sum(action_method_counts.get(m, 0) for m in manager_supported_methods)

# --- Goal cycle analysis ---
total_goal_dates = df['总目标达成时间'].dropna().dt.to_period('M').value_counts().sort_index().reset_index()
total_goal_dates.columns = ['月份', '人数']
total_goal_dates['月份'] = total_goal_dates['月份'].astype(str)

mid_dates = df['中期回顾时间'].dropna().dt.to_period('M').value_counts().sort_index().reset_index()
mid_dates.columns = ['月份', '人数']
mid_dates['月份'] = mid_dates['月份'].astype(str)

end_dates = df['终期回顾时间'].dropna().dt.to_period('M').value_counts().sort_index().reset_index()
end_dates.columns = ['月份', '人数']
end_dates['月份'] = end_dates['月份'].astype(str)

cycle_length_stats = pd.DataFrame({
    '指标': ['填写到总目标天数', '中期到终期天数', '终期到总目标天数'],
    '平均天数': [
        df['填写到总目标天数'].mean(),
        df['中期到终期天数'].mean(),
        df['终期到总目标天数'].mean()
    ],
    '中位数天数': [
        df['填写到总目标天数'].median(),
        df['中期到终期天数'].median(),
        df['终期到总目标天数'].median()
    ],
    '最小天数': [
        df['填写到总目标天数'].min(),
        df['中期到终期天数'].min(),
        df['终期到总目标天数'].min()
    ],
    '最大天数': [
        df['填写到总目标天数'].max(),
        df['中期到终期天数'].max(),
        df['终期到总目标天数'].max()
    ]
})

anomalies = pd.DataFrame({
    '异常类型': ['总目标达成时间早于填写日期', '中期回顾时间晚于终期回顾时间', '终期回顾时间晚于总目标达成时间'],
    '异常记录数': [
        df['异常_总目标早于填写'].sum(),
        df['异常_中期晚于终期'].sum(),
        df['异常_终期晚于总目标'].sum()
    ]
})

# ============================================================
# 5. Generate Charts
# ============================================================
print("Generating charts...")

def save_chart(fig, filename):
    path = os.path.join(CHART_DIR, filename)
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return path

# Chart 1: Position distribution
fig, ax = plt.subplots(figsize=(8, 5))
ax.bar(position_dist['目前职位'], position_dist['人数'], color='steelblue')
ax.set_title('目前职位分布')
ax.set_xlabel('职位')
ax.set_ylabel('人数')
plt.xticks(rotation=30, ha='right')
chart_position = save_chart(fig, 'position_dist.png')

# Chart 2: Goal theme distribution
fig, ax = plt.subplots(figsize=(8, 5))
ax.barh(theme_dist['主题'], theme_dist['出现次数'], color='coral')
ax.set_title('发展目标主题分布')
ax.set_xlabel('出现次数')
chart_theme = save_chart(fig, 'theme_dist.png')

# Chart 3: Skill attribute distribution
fig, ax = plt.subplots(figsize=(6, 6))
ax.pie(skill_attr_dist['出现次数'], labels=skill_attr_dist['技能属性'], autopct='%1.1f%%', startangle=90)
ax.set_title('技能属性分布（软性 vs 硬性）')
chart_skill = save_chart(fig, 'skill_attr_dist.png')

# Chart 4: Action coverage
fig, ax = plt.subplots(figsize=(8, 5))
x = np.arange(len(action_coverage))
width = 0.35
ax.bar(x - width/2, action_coverage['目标1填写数'], width, label='目标1', color='teal')
ax.bar(x + width/2, action_coverage['目标2填写数'], width, label='目标2', color='orange')
ax.set_xticks(x)
ax.set_xticklabels(action_coverage['维度'])
ax.set_ylabel('填写人数')
ax.set_title('70-20-10 行动计划填写情况')
ax.legend()
chart_action = save_chart(fig, 'action_coverage.png')

# Chart 5: Completeness distribution
fig, ax = plt.subplots(figsize=(6, 6))
ax.pie(completeness_dist['人数'], labels=completeness_dist['完整度等级'], autopct='%1.1f%%', startangle=90)
ax.set_title('IDP 完整度等级分布')
chart_completeness = save_chart(fig, 'completeness_dist.png')

# Chart 6: Total goal date timeline
fig, ax = plt.subplots(figsize=(10, 5))
ax.plot(total_goal_dates['月份'].astype(str), total_goal_dates['人数'], marker='o', color='purple')
ax.set_title('总目标达成时间分布（按月）')
ax.set_xlabel('月份')
ax.set_ylabel('人数')
plt.xticks(rotation=45, ha='right')
chart_timeline = save_chart(fig, 'total_goal_timeline.png')

# ============================================================
# 6. Generate Excel Workbook
# ============================================================
print("Generating Excel workbook...")
excel_path = os.path.join(OUTPUT_DIR, '个人发展计划表分析_结果.xlsx')

# Prepare raw data sheet with clean columns
raw_cols = ['工号', '员工姓名', '经理姓名', '目前职位', '店铺', '填写日期',
            '目标1类型', '目标1描述', '目标1_70计划', '目标1_70描述', '目标1_20计划', '目标1_20描述', '目标1_10计划', '目标1_10描述', '目标1资源',
            '目标2类型', '目标2描述', '目标2_70计划', '目标2_70描述', '目标2_20计划', '目标2_20描述', '目标2_10计划', '目标2_10描述', '目标2资源',
            '接受调动', '调动目标', '总目标达成时间', '中期回顾时间', '终期回顾时间',
            '目标总数', '完整度得分', '完整度百分比', '完整度等级', '目标填写情况']
df_raw = df[raw_cols].copy()

# Prepare employee score sheet
score_cols = ['工号', '员工姓名', '经理姓名', '目前职位', '店铺', '目标总数', '完整度得分', '完整度百分比', '完整度等级',
              '目标1已填', '目标1_70已填', '目标1_20已填', '目标1_10已填', '目标1资源已填',
              '目标2已填', '目标2_70已填', '目标2_20已填', '目标2_10已填', '目标2资源已填',
              '填写到总目标天数', '中期到终期天数', '终期到总目标天数']
df_score = df[score_cols].copy()

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    # Raw data
    df_raw.to_excel(writer, sheet_name='原始数据', index=False)
    
    # Basic situation
    position_dist.to_excel(writer, sheet_name='基本情况汇总', startrow=0, startcol=0, index=False)
    manager_dist.to_excel(writer, sheet_name='基本情况汇总', startrow=0, startcol=4, index=False)
    store_dist.to_excel(writer, sheet_name='基本情况汇总', startrow=0, startcol=8, index=False)
    mobility_dist.to_excel(writer, sheet_name='基本情况汇总', startrow=0, startcol=12, index=False)
    fill_date_dist.to_excel(writer, sheet_name='基本情况汇总', startrow=0, startcol=16, index=False)
    
    # Goal analysis
    goal_completeness_dist.to_excel(writer, sheet_name='目标分析', startrow=0, startcol=0, index=False)
    goal_type_combined.to_excel(writer, sheet_name='目标分析', startrow=0, startcol=4, index=False)
    skill_attr_dist.to_excel(writer, sheet_name='目标分析', startrow=0, startcol=8, index=False)
    theme_dist.to_excel(writer, sheet_name='目标分析', startrow=0, startcol=12, index=False)
    keyword_df.to_excel(writer, sheet_name='目标分析', startrow=0, startcol=16, index=False)
    
    # Action plan analysis
    action_coverage.to_excel(writer, sheet_name='行动计划分析', startrow=0, startcol=0, index=False)
    action_method_df.to_excel(writer, sheet_name='行动计划分析', startrow=0, startcol=7, index=False)
    resource_keyword_df.to_excel(writer, sheet_name='行动计划分析', startrow=0, startcol=12, index=False)
    completeness_dist.to_excel(writer, sheet_name='行动计划分析', startrow=0, startcol=18, index=False)
    
    # Cycle analysis
    total_goal_dates.to_excel(writer, sheet_name='周期分析', startrow=0, startcol=0, index=False)
    mid_dates.to_excel(writer, sheet_name='周期分析', startrow=0, startcol=4, index=False)
    end_dates.to_excel(writer, sheet_name='周期分析', startrow=0, startcol=8, index=False)
    cycle_length_stats.to_excel(writer, sheet_name='周期分析', startrow=0, startcol=12, index=False)
    anomalies.to_excel(writer, sheet_name='周期分析', startrow=0, startcol=19, index=False)
    
    # Employee score
    df_score.to_excel(writer, sheet_name='员工明细评分', index=False)

# Format Excel workbook
wb = openpyxl.load_workbook(excel_path)
header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
header_font = Font(bold=True)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format header row
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

wb.save(excel_path)
print(f"Excel saved: {excel_path}")

# ============================================================
# 7. Generate Markdown Report
# ============================================================
print("Generating Markdown report...")
report_path = os.path.join(OUTPUT_DIR, '个人发展计划表分析_报告.md')

report_lines = []
report_lines.append("# 个人发展计划表（IDP）数据分析报告")
report_lines.append("")
report_lines.append(f"**分析日期：** {datetime.now().strftime('%Y年%m月%d日')}")
report_lines.append(f"**数据来源：** 1. Raw/个人发展计划表ID.xlsx")
report_lines.append(f"**分析样本：** {num_employees} 名员工")
report_lines.append("")

# Executive summary
report_lines.append("## 一、执行摘要")
report_lines.append("")
report_lines.append(f"- 本次分析覆盖 **{num_employees} 名**员工，来自 **{num_stores} 家**店铺，由 **{num_managers} 位**经理管理。")
position_summary = '，'.join([f"{row['目前职位']} {row['人数']}人" for _, row in position_dist.iterrows()])
report_lines.append(f"- 职位分布：{position_summary}。")
report_lines.append(f"- 目标填写完整度：{df['目标总数'].sum()} 个目标被填写，人均 {(df['目标总数'].sum() / num_employees):.2f} 个目标。")
report_lines.append(f"- IDP 平均完整度得分：{df['完整度得分'].mean():.1f} / 12（{df['完整度百分比'].mean():.1f}%）。")
report_lines.append(f"- 高完整度员工：{(df['完整度等级'] == '高').sum()} 人；中完整度：{(df['完整度等级'] == '中').sum()} 人；低完整度：{(df['完整度等级'] == '低').sum()} 人。")
report_lines.append("")

# Basic situation
report_lines.append("## 二、基本情况分析")
report_lines.append("")
report_lines.append("### 2.1 人员与组织分布")
report_lines.append("")
report_lines.append(f"- **员工总数**：{num_employees}")
report_lines.append(f"- **经理人数**：{num_managers}")
report_lines.append(f"- **店铺数量**：{num_stores}")
report_lines.append(f"- **职位类型数**：{num_positions}")
report_lines.append("")
report_lines.append("### 2.2 职位分布")
report_lines.append("")
report_lines.append(position_dist.to_markdown(index=False))
report_lines.append("")
report_lines.append(f"![职位分布](charts/position_dist.png)")
report_lines.append("")

report_lines.append("### 2.3 经理下属分布（Top 10）")
report_lines.append("")
report_lines.append(manager_dist.head(10).to_markdown(index=False))
report_lines.append("")

report_lines.append("### 2.4 店铺分布（Top 10）")
report_lines.append("")
report_lines.append(store_dist.head(10).to_markdown(index=False))
report_lines.append("")

report_lines.append("### 2.5 跨城市/跨部门调动意愿")
report_lines.append("")
report_lines.append(mobility_dist.to_markdown(index=False))
report_lines.append("")

# IDP goal analysis
report_lines.append("## 三、IDP 目标分析")
report_lines.append("")
report_lines.append("### 3.1 目标填写完整度")
report_lines.append("")
report_lines.append(goal_completeness_dist.to_markdown(index=False))
report_lines.append("")

report_lines.append("### 3.2 目标类型分布（综合目标1与目标2）")
report_lines.append("")
report_lines.append(goal_type_combined.to_markdown(index=False))
report_lines.append("")

report_lines.append("### 3.3 技能属性分布（软性 vs 硬性）")
report_lines.append("")
report_lines.append(skill_attr_dist.to_markdown(index=False))
report_lines.append("")
report_lines.append(f"![技能属性分布](charts/skill_attr_dist.png)")
report_lines.append("")

report_lines.append("### 3.4 目标主题分布")
report_lines.append("")
report_lines.append(theme_dist.to_markdown(index=False))
report_lines.append("")
report_lines.append(f"![目标主题分布](charts/theme_dist.png)")
report_lines.append("")

report_lines.append("### 3.5 目标描述高频关键词（Top 20）")
report_lines.append("")
report_lines.append(keyword_df.head(20).to_markdown(index=False))
report_lines.append("")

# Action plan analysis
report_lines.append("## 四、行动计划分析")
report_lines.append("")
report_lines.append("### 4.1 70-20-10 学习模型填写率")
report_lines.append("")
report_lines.append(action_coverage.to_markdown(index=False))
report_lines.append("")
report_lines.append(f"![行动计划填写情况](charts/action_coverage.png)")
report_lines.append("")

report_lines.append("### 4.2 常见行动方式分布")
report_lines.append("")
report_lines.append(action_method_df.to_markdown(index=False))
report_lines.append("")

report_lines.append("### 4.3 资源需求分析")
report_lines.append("")
report_lines.append(f"- 资源字段填写率：{resource_filled_rate:.1f}%")
report_lines.append("- 高频资源关键词：")
report_lines.append(resource_keyword_df.to_markdown(index=False))
report_lines.append("")

report_lines.append("### 4.4 IDP 完整度评分分布")
report_lines.append("")
report_lines.append(completeness_dist.to_markdown(index=False))
report_lines.append("")
report_lines.append(f"![完整度等级分布](charts/completeness_dist.png)")
report_lines.append("")

report_lines.append("### 4.5 经理支持度分析")
report_lines.append("")
report_lines.append(f"- 依赖经理支持的学习方式（辅导、向经理寻求反馈、人脉交流学习）共出现 **{manager_support_count} 次**。")
report_lines.append(f"- 占所有行动方式记录的 **{manager_support_count / sum(action_method_counts.values()) * 100:.1f}%**。")
report_lines.append("")

# Goal cycle analysis
report_lines.append("## 五、目标周期分析")
report_lines.append("")
report_lines.append("### 5.1 周期长度统计")
report_lines.append("")
report_lines.append(cycle_length_stats.to_markdown(index=False))
report_lines.append("")

report_lines.append("### 5.2 总目标达成时间分布（按月）")
report_lines.append("")
report_lines.append(total_goal_dates.to_markdown(index=False))
report_lines.append("")
report_lines.append(f"![总目标达成时间分布](charts/total_goal_timeline.png)")
report_lines.append("")

report_lines.append("### 5.3 时间异常检查")
report_lines.append("")
report_lines.append(anomalies.to_markdown(index=False))
report_lines.append("")

# Insights and recommendations
report_lines.append("## 六、洞察与建议")
report_lines.append("")
report_lines.append("### 6.1 基本情况洞察")
report_lines.append("")
report_lines.append(f"- 经理覆盖较{'均匀' if manager_dist['下属人数'].std() < 3 else '不均衡'}，人均下属 {manager_dist['下属人数'].mean():.1f} 人，标准差 {manager_dist['下属人数'].std():.1f}。")
report_lines.append(f"- 调动意愿：{(df['接受调动'] == '是').sum()} 人愿意接受调动，{(df['接受调动'] == '否').sum()} 人不愿意，其余未明确。")
report_lines.append("")
report_lines.append("### 6.2 IDP 目标洞察")
report_lines.append("")
# Top theme (skip 未填写)
theme_dist_valid = theme_dist[theme_dist['主题'] != '未填写']
if not theme_dist_valid.empty:
    top_theme = theme_dist_valid.iloc[0]['主题']
    top_theme_count = theme_dist_valid.iloc[0]['出现次数']
    report_lines.append(f"- 最受欢迎的发展主题是 **{top_theme}**（{top_theme_count} 次），建议针对该主题设计集中培训或工作坊。")
else:
    report_lines.append("- 暂无有效发展目标主题数据。")
# Soft vs hard ratio
soft_count = skill_attr_dist[skill_attr_dist['技能属性'] == '软性技能']['出现次数'].sum()
hard_count = skill_attr_dist[skill_attr_dist['技能属性'] == '硬性技能']['出现次数'].sum()
report_lines.append(f"- 软性技能目标占比 {soft_count / (soft_count + hard_count) * 100:.1f}%，硬性技能占比 {hard_count / (soft_count + hard_count) * 100:.1f}%。")
report_lines.append(f"- 仅填写一个目标或未填写目标的员工有 {((df['目标总数'] == 0) | (df['目标总数'] == 1)).sum()} 人，建议推动全员填写双目标。")
report_lines.append("")
report_lines.append("### 6.3 行动计划洞察")
report_lines.append("")
report_lines.append(f"- 70% 工作中实践填写率最高（目标1：{df['目标1_70已填'].sum() / len(df) * 100:.1f}%，目标2：{df['目标2_70已填'].sum() / len(df) * 100:.1f}%），符合 IDP 设计原则。")
report_lines.append(f"- 10% 正式课程填写率相对较低（目标1：{df['目标1_10已填'].sum() / len(df) * 100:.1f}%，目标2：{df['目标2_10已填'].sum() / len(df) * 100:.1f}%），可加强课程推荐与培训资源对接。")
report_lines.append(f"- 平均完整度得分 {df['完整度得分'].mean():.1f}/12，仍有提升空间。建议对完整度低于 6 分的员工进行专项跟进。")
report_lines.append("")
report_lines.append("### 6.4 目标周期洞察")
report_lines.append("")
report_lines.append(f"- 平均目标周期为 {df['填写到总目标天数'].mean():.0f} 天，中期到终期平均 {df['中期到终期天数'].mean():.0f} 天。")
report_lines.append(f"- 发现时间异常记录 {anomalies['异常记录数'].sum()} 条，建议与相关经理核对并修正。")
report_lines.append("")
report_lines.append("### 6.5 行动建议")
report_lines.append("")
report_lines.append("1. **推动全员双目标**：针对仅填写一个目标或未填写目标的员工，由经理在 1 对 1 中补充第二目标。")
report_lines.append("2. **强化课程资源对接**：10% 正式课程维度填写率偏低，建议 HR/L&D 整理课程清单并定向推荐。")
report_lines.append("3. **建立经理支持机制**：鼓励经理主动提供反馈、辅导和人脉资源，提升 20% 向他人学习维度的质量。")
report_lines.append("4. **修正时间异常**：对时间倒置的记录进行复核，确保中期回顾、终期回顾与总目标时间逻辑一致。")
report_lines.append("5. **定期跟踪完整度**：将 IDP 完整度作为季度管理指标，对低完整度员工及时提醒。")
report_lines.append("")

# Appendix
report_lines.append("## 七、附录")
report_lines.append("")
report_lines.append("### 7.1 数据字典")
report_lines.append("")
report_lines.append("- 完整度得分：每位员工的目标类型、目标描述、三个行动计划维度（70/20/10）、资源字段共 12 个字段的填写情况。")
report_lines.append("- 70-20-10 模型：70% 工作中实践、20% 向他人学习、10% 正式课程。")
report_lines.append("- 软性技能：沟通、计划与组织、结果导向等；硬性技能：销售艺术、卓越客户体验、品牌与文化传承等。")
report_lines.append("")
report_lines.append("### 7.2 分析方法")
report_lines.append("")
report_lines.append("- 使用 Python pandas 进行数据清洗与统计。")
report_lines.append("- 使用 jieba 进行中文关键词提取。")
report_lines.append("- 使用 matplotlib 生成图表。")
report_lines.append("- 使用 openpyxl 生成 Excel 分析表。")
report_lines.append("")

with open(report_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(report_lines))

print(f"Report saved: {report_path}")
print("All outputs generated successfully.")
