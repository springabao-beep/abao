import os
import sys
import argparse

# Parse arguments before importing matplotlib
parser = argparse.ArgumentParser(description='Analyze IDP Excel file')
parser.add_argument('input_file', help='Path to IDP Excel file')
parser.add_argument('output_dir', nargs='?', help='Output directory')
args = parser.parse_args()

INPUT_FILE = args.input_file
if args.output_dir:
    OUTPUT_DIR = args.output_dir
else:
    OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(INPUT_FILE)), 'Final')

mpl_config_dir = os.path.join(OUTPUT_DIR, '.mpl_config')
os.environ['MPLCONFIGDIR'] = mpl_config_dir
os.makedirs(os.environ['MPLCONFIGDIR'], exist_ok=True)

CHART_DIR = os.path.join(OUTPUT_DIR, "charts")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(CHART_DIR, exist_ok=True)

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib import font_manager as fm_manager
import jieba
import jieba.analyse
import re
from collections import Counter
from datetime import datetime

sys.stdout.reconfigure(encoding='utf-8')

# Matplotlib font setup
chinese_font = None
font_candidates = ['Microsoft YaHei', 'SimHei', 'SimSun', 'DengXian', 'Arial']
for candidate in font_candidates:
    try:
        font_path = fm_manager.findfont(candidate, fallback_to_default=False)
        if font_path and candidate.lower() in font_path.lower():
            chinese_font = candidate
            break
    except Exception:
        continue

if chinese_font is None:
    system_fonts = fm_manager.findSystemFonts(fontpaths=None, fontext='ttf')
    for fp in system_fonts:
        if any(k in fp.lower() for k in ['yahei', 'simhei', 'simsun', 'arial']):
            chinese_font = fm.FontProperties(fname=fp).get_name()
            break

if chinese_font:
    plt.rcParams['font.sans-serif'] = [chinese_font]
plt.rcParams['axes.unicode_minus'] = False

# ============================================================
# Read & Clean Data
# ============================================================
print("Reading Excel file...")
df = pd.read_excel(INPUT_FILE, sheet_name=0, header=2)

# Drop rows where employee ID is missing
df = df[df['工号'].notna()].copy()

# Standardize column names to English
column_map = {
    '工号': 'Employee_ID',
    '员工姓名': 'Employee_Name',
    '经理姓名': 'Manager_Name',
    '目前职位': 'Position',
    '区域': 'Region',
    '店铺': 'Store',
    '日期': 'Fill_Date',
    '请勾选发展目标1': 'Goal1_Type',
    '目标1的详细描述': 'Goal1_Description',
    '目标1的行动计划（70%-工作中实践）': 'Goal1_70_Plan',
    '目标1的行动描述（70%-工作中实践）': 'Goal1_70_Detail',
    '目标1的行动计划（20%-向他人学习）': 'Goal1_20_Plan',
    '目标1的行动描述（20%-向他人学习）': 'Goal1_20_Detail',
    '目标1的行动计划（10%-正式课程）': 'Goal1_10_Plan',
    '目标1的行动描述（10%-正式课程）': 'Goal1_10_Detail',
    '目标1的其它所需资源': 'Goal1_Resources',
    '请勾选发展目标2': 'Goal2_Type',
    '目标2的详细描述': 'Goal2_Description',
    '目标2的行动计划（70%-工作中实践）': 'Goal2_70_Plan',
    '目标2的行动描述（70%-工作中实践）': 'Goal2_70_Detail',
    '目标2的行动计划（20%-向他人学习）': 'Goal2_20_Plan',
    '目标2的行动描述（20%-向他人学习）': 'Goal2_20_Detail',
    '目标2的行动计划（10%-正式课程）': 'Goal2_10_Plan',
    '目标2的行动描述（10%-正式课程）': 'Goal2_10_Detail',
    '目标2的其它需要资源': 'Goal2_Resources',
    '您是否接受跨城市/跨部门的调动？': 'Mobility_Accept',
    '请写明您想调动至哪里？': 'Mobility_Target',
    '总目标达成时间': 'Total_Goal_Date',
    '中期回顾时间': 'Mid_Review_Date',
    '终期回顾时间': 'Final_Review_Date'
}

df.rename(columns=column_map, inplace=True)

# Trim whitespace on text columns
text_cols = ['Employee_Name', 'Manager_Name', 'Position', 'Region', 'Store',
             'Goal1_Type', 'Goal2_Type', 'Mobility_Accept']
for col in text_cols:
    if col in df.columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace('nan', np.nan)
        df[col] = df[col].replace('', np.nan)

# Normalize position names
def normalize_position(pos):
    if pd.isna(pos):
        return pos
    p = str(pos).strip().lower().replace(' ', '')
    if p in ['ssa', 'seniorsalesassociate']:
        return 'SSA'
    elif p in ['sa', 'sales', 'seniorsales', 'salesassociate']:
        return 'SA'
    elif p in ['supervisor', 'supvisor', 'spv']:
        return 'Supervisor'
    elif p in ['admin', 'administrator']:
        return 'Admin'
    elif p in ['senioradmin', 'senioradministrator']:
        return 'Senior Admin'
    elif p in ['eliteclientsupervisor']:
        return 'Elite Client Supervisor'
    else:
        return str(pos).strip()

df['Position'] = df['Position'].apply(normalize_position)

# Parse dates
def parse_chinese_date(x):
    if pd.isna(x) or x == 'nan':
        return pd.NaT
    if isinstance(x, pd.Timestamp):
        return x
    try:
        s = str(x).strip()
        s = s.replace('年', '-').replace('月', '-').replace('日', '').strip('-')
        return pd.to_datetime(s, errors='coerce')
    except:
        return pd.NaT

for col in ['Fill_Date', 'Total_Goal_Date', 'Mid_Review_Date', 'Final_Review_Date']:
    df[col] = df[col].apply(parse_chinese_date)

# Fill missing goal types
df['Goal1_Type'] = df['Goal1_Type'].replace('', np.nan)
df['Goal2_Type'] = df['Goal2_Type'].replace('', np.nan)

# Translate mobility values to English
df['Mobility_Accept'] = df['Mobility_Accept'].replace({'是': 'Yes', '否': 'No'})

print(f"Loaded {len(df)} employee records.")

# ============================================================
# Derived Fields
# ============================================================
df['Goal1_Filled'] = df['Goal1_Type'].notna()
df['Goal2_Filled'] = df['Goal2_Type'].notna()
df['Goals_Total'] = df['Goal1_Filled'].astype(int) + df['Goal2_Filled'].astype(int)

# Skill type classification
def extract_skill_type(goal_type):
    if pd.isna(goal_type):
        return 'Not Filled'
    if '软性' in goal_type:
        return 'Soft Skill'
    elif '硬性' in goal_type:
        return 'Hard Skill'
    else:
        return 'Other'

df['Goal1_Skill_Type'] = df['Goal1_Type'].apply(extract_skill_type)
df['Goal2_Skill_Type'] = df['Goal2_Type'].apply(extract_skill_type)

# Goal theme classification
def classify_goal_theme(goal_type):
    if pd.isna(goal_type):
        return 'Not Filled'
    gt = goal_type
    if '沟通' in gt:
        return 'Communication'
    elif '计划' in gt or '组织' in gt:
        return 'Planning & Organization'
    elif '销售' in gt:
        return 'Sales Art'
    elif '客户' in gt or '卓越' in gt:
        return 'Customer Excellence'
    elif '品牌' in gt or '文化' in gt:
        return 'Brand & Culture'
    elif '结果' in gt or '结果导向' in gt:
        return 'Result Orientation'
    elif '领导' in gt or '团队' in gt:
        return 'Leadership / Team'
    elif '数据分析' in gt or '数据' in gt:
        return 'Data Analysis'
    elif '管理' in gt:
        return 'Management'
    else:
        return 'Other'

df['Goal1_Theme'] = df['Goal1_Type'].apply(classify_goal_theme)
df['Goal2_Theme'] = df['Goal2_Type'].apply(classify_goal_theme)

# Action plan coverage (70/20/10)
def action_filled(plan, desc):
    return (pd.notna(plan) and str(plan).strip() != '') or (pd.notna(desc) and str(desc).strip() != '')

for goal in ['Goal1', 'Goal2']:
    for pct in ['70', '20', '10']:
        plan_col = f'{goal}_{pct}_Plan'
        desc_col = f'{goal}_{pct}_Detail'
        df[f'{goal}_{pct}_Filled'] = df.apply(lambda row: action_filled(row.get(plan_col), row.get(desc_col)), axis=1)

# Action methods extraction
action_method_keywords = {
    'Assist with new tasks/projects': ['协助新任务', '协助任务'],
    'Lead new tasks/projects': ['负责新任务', '负责项目'],
    'Assist with struggling tasks': ['协助解决', '陷入困境'],
    'Host store meetings/showcase results': ['主持店铺会议', '展示团队', '展示成果', '展示成果'],
    'Attend cross-team/store meetings': ['参加跨小组', '跨团队', '跨店铺', '职能会议'],
    'Evaluate & optimize processes': ['评估并优化', '优化现有流程'],
    'Observe excellent colleagues': ['观察优秀同事', '观察优秀'],
    'Coaching/Mentoring': ['辅导'],
    'Seek feedback from manager': ['向经理寻求反馈', '经理反馈'],
    'Networking & peer learning': ['人脉交流', '交流学习'],
    'Online learning (Harvard platform)': ['线上学习', '哈佛管理'],
    'Offline training (workshops)': ['线下培训', '公司培训', '工作坊'],
    'Self-directed learning (books, podcasts, languages)': ['自主学习', '阅读书籍', '收听播客', '学习语言']
}

def extract_action_methods(text):
    if pd.isna(text):
        return []
    text_str = str(text)
    found = []
    for method, keywords in action_method_keywords.items():
        if any(kw in text_str for kw in keywords):
            found.append(method)
    return found

all_action_methods = []
for goal in ['Goal1', 'Goal2']:
    for pct in ['70', '20', '10']:
        col = f'{goal}_{pct}_Plan'
        df[f'{goal}_{pct}_Methods'] = df[col].apply(extract_action_methods)
        all_action_methods.extend(df[f'{goal}_{pct}_Methods'].tolist())

action_method_list = []
for methods in all_action_methods:
    action_method_list.extend(methods)
action_method_counts = Counter(action_method_list)

# Resource columns filled
df['Goal1_Res_Filled'] = df['Goal1_Resources'].apply(lambda x: pd.notna(x) and str(x).strip() != '' and str(x).strip() != '无')
df['Goal2_Res_Filled'] = df['Goal2_Resources'].apply(lambda x: pd.notna(x) and str(x).strip() != '' and str(x).strip() != '无')

# Completeness score (max 12)
df['Goal1_Desc_Filled'] = df['Goal1_Description'].apply(lambda x: pd.notna(x) and str(x).strip() != '')
df['Goal2_Desc_Filled'] = df['Goal2_Description'].apply(lambda x: pd.notna(x) and str(x).strip() != '')

df['Completeness_Score'] = (
    df['Goal1_Filled'].astype(int) +
    df['Goal1_Desc_Filled'].astype(int) +
    df['Goal1_70_Filled'].astype(int) +
    df['Goal1_20_Filled'].astype(int) +
    df['Goal1_10_Filled'].astype(int) +
    df['Goal1_Res_Filled'].astype(int) +
    df['Goal2_Filled'].astype(int) +
    df['Goal2_Desc_Filled'].astype(int) +
    df['Goal2_70_Filled'].astype(int) +
    df['Goal2_20_Filled'].astype(int) +
    df['Goal2_10_Filled'].astype(int) +
    df['Goal2_Res_Filled'].astype(int)
)

df['Completeness_Pct'] = (df['Completeness_Score'] / 12 * 100).round(1)

def completeness_tier(score):
    if score >= 10:
        return 'High'
    elif score >= 6:
        return 'Medium'
    else:
        return 'Low'

df['Completeness_Tier'] = df['Completeness_Score'].apply(completeness_tier)

# Cycle analysis
df['Days_Fill_to_Total'] = (df['Total_Goal_Date'] - df['Fill_Date']).dt.days
df['Days_Mid_to_Final'] = (df['Final_Review_Date'] - df['Mid_Review_Date']).dt.days
df['Days_Final_to_Total'] = (df['Total_Goal_Date'] - df['Final_Review_Date']).dt.days

# Anomaly flags
df['Anomaly_Total_Before_Fill'] = df['Days_Fill_to_Total'] < 0
df['Anomaly_Mid_After_Final'] = df['Days_Mid_to_Final'] < 0
df['Anomaly_Final_After_Total'] = df['Days_Final_to_Total'] < 0

# ============================================================
# Analysis Summaries
# ============================================================
print("Computing summaries...")

# --- Basic situation ---
num_employees = len(df)
num_managers = df['Manager_Name'].nunique()
num_stores = df['Store'].nunique()
num_positions = df['Position'].nunique()
num_regions = df['Region'].nunique() if 'Region' in df.columns else 0

position_dist = df['Position'].value_counts().reset_index()
position_dist.columns = ['Position', 'Count']

manager_dist = df['Manager_Name'].value_counts().reset_index()
manager_dist.columns = ['Manager', 'Direct_Reports']

store_dist = df['Store'].value_counts().reset_index()
store_dist.columns = ['Store', 'Count']

if 'Region' in df.columns:
    region_dist = df['Region'].value_counts().reset_index()
    region_dist.columns = ['Region', 'Count']

mobility_dist = df['Mobility_Accept'].value_counts(dropna=False).reset_index()
mobility_dist.columns = ['Mobility_Accept', 'Count']

# --- IDP goal analysis ---
goal1_types = df['Goal1_Type'].value_counts(dropna=False).reset_index()
goal1_types.columns = ['Goal_Type', 'Count']
goal1_types['Goal_Type'] = goal1_types['Goal_Type'].fillna('Not Filled')

goal2_types = df['Goal2_Type'].value_counts(dropna=False).reset_index()
goal2_types.columns = ['Goal_Type', 'Count']
goal2_types['Goal_Type'] = goal2_types['Goal_Type'].fillna('Not Filled')

all_goals = pd.concat([
    df['Goal1_Type'].rename('Goal_Type'),
    df['Goal2_Type'].rename('Goal_Type')
])
all_goals = all_goals[all_goals.notna()]
goal_type_combined = all_goals.value_counts().reset_index()
goal_type_combined.columns = ['Goal_Type', 'Count']

# Skill attribute distribution
skill_attr_dist = pd.concat([
    df['Goal1_Skill_Type'].rename('Skill_Type'),
    df['Goal2_Skill_Type'].rename('Skill_Type')
]).value_counts().reset_index()
skill_attr_dist.columns = ['Skill_Type', 'Count']

# Theme distribution
theme_dist = pd.concat([
    df['Goal1_Theme'].rename('Theme'),
    df['Goal2_Theme'].rename('Theme')
]).value_counts().reset_index()
theme_dist.columns = ['Theme', 'Count']

# Goal completeness categories
def goal_completeness_category(row):
    g1 = row['Goal1_Filled']
    g2 = row['Goal2_Filled']
    if g1 and g2:
        return 'Both Goals Filled'
    elif g1 and not g2:
        return 'Only Goal 1 Filled'
    elif not g1 and g2:
        return 'Only Goal 2 Filled'
    else:
        return 'No Goals Filled'

df['Goal_Fill_Status'] = df.apply(goal_completeness_category, axis=1)
goal_completeness_dist = df['Goal_Fill_Status'].value_counts().reset_index()
goal_completeness_dist.columns = ['Goal_Fill_Status', 'Count']

# Keyword extraction using jieba
stopwords = set(['的', '了', '在', '是', '和', '与', '及', '等', '对', '为', '有', '我', '要', '将', '并', '以', '以及',
    '通过', '提升', '学习', '能力', '工作', '目标', '计划', '行动', '提高', '加强', '优化', '进行', '需要', '可以', '能够',
    '希望', '做到', '达到', '完成', '实现', '帮助', '协助', '负责', '每周', '每日', '每月', '每季度', '一次', '一个', '一些',
    '不同', '相关', '有关', '对于', '关于', '根据', '按照', '结合', '针对', '不断', '持续', '进一步', '更好', '更多', '更加',
    '有效', '高效', '良好', '积极', '主动', '及时', '定期', '充分', '全面', '深入', '系统', '专业', '实际', '具体', '明确',
    '清晰', '合理', '科学', '规范', '标准', '流程', '方法', '方式', '措施', '方案', '路径', '渠道', '平台', '资源', '支持',
    '反馈', '指导', '沟通', '交流', '分享', '汇报', '总结', '复盘', '反思', '改进', '改善', '完善', '落实', '执行', '推进',
    '推动', '促进', '带动', '确保', '保障', '维护', '管理', '经营', '运营', '服务', '销售', '客户', '团队', '员工', '同事',
    '经理', '领导', '店铺', '公司', '集团', '部门', '区域', '门店', '柜台', '业务', '项目', '任务', '活动', '会议', '课程',
    '培训', '书籍', '播客', '语言', '无', '其他', '做好'])

desc_text = ' '.join(df['Goal1_Description'].fillna('').astype(str)) + ' ' + ' '.join(df['Goal2_Description'].fillna('').astype(str))
words = jieba.lcut(desc_text)
words = [w for w in words if len(w) >= 2 and w not in stopwords and not re.match(r'^[a-zA-Z0-9]+$', w)]
keyword_counts = Counter(words).most_common(30)
keyword_df = pd.DataFrame(keyword_counts, columns=['Keyword', 'Frequency'])
keyword_df['Keyword_EN'] = keyword_df['Keyword'].map({
    '沟通': 'Communication', '销售': 'Sales', '客户': 'Customer', '团队': 'Team',
    '品牌': 'Brand', '文化': 'Culture', '计划': 'Planning', '组织': 'Organization',
    '结果': 'Results', '领导': 'Leadership', '数据': 'Data', '分析': 'Analysis',
    '管理': 'Management', '服务': 'Service', '体验': 'Experience', '接待': 'Reception',
    '邀约': 'Invitation', '转化': 'Conversion', '复购': 'Repurchase',
    '业绩': 'Performance', '话术': 'Script', '复盘': 'Review',
    '班组长': 'Team Leader', '流程': 'Process', '指标': 'KPI',
    '带教': 'Coaching', '新员工': 'New Employee', '专业知识': 'Professional Knowledge',
    '表达': 'Expression', '统筹': 'Coordination', '执行': 'Execution',
    '汇报': 'Reporting', '排期': 'Scheduling', '风险': 'Risk',
    '盘点': 'Inventory', '调价': 'Price Change', '营运': 'Operations',
    '合规': 'Compliance', '邀约话术': 'Invitation Script', '转化率': 'Conversion Rate',
    '客单价': 'Average Transaction Value', '复购率': 'Repurchase Rate'
}).fillna(keyword_df['Keyword'])

# --- Action plan analysis ---
action_coverage = pd.DataFrame({
    'Dimension': ['70% On-the-job Practice', '20% Learning from Others', '10% Formal Courses'],
    'Goal1_Filled': [
        int(df['Goal1_70_Filled'].sum()),
        int(df['Goal1_20_Filled'].sum()),
        int(df['Goal1_10_Filled'].sum())
    ],
    'Goal1_Rate': [
        f"{df['Goal1_70_Filled'].sum() / len(df) * 100:.1f}%",
        f"{df['Goal1_20_Filled'].sum() / len(df) * 100:.1f}%",
        f"{df['Goal1_10_Filled'].sum() / len(df) * 100:.1f}%"
    ],
    'Goal2_Filled': [
        int(df['Goal2_70_Filled'].sum()),
        int(df['Goal2_20_Filled'].sum()),
        int(df['Goal2_10_Filled'].sum())
    ],
    'Goal2_Rate': [
        f"{df['Goal2_70_Filled'].sum() / len(df) * 100:.1f}%",
        f"{df['Goal2_20_Filled'].sum() / len(df) * 100:.1f}%",
        f"{df['Goal2_10_Filled'].sum() / len(df) * 100:.1f}%"
    ]
})

action_method_df = pd.DataFrame(action_method_counts.most_common(), columns=['Action_Method', 'Count'])

resource_filled_rate = ((df['Goal1_Res_Filled'].sum() + df['Goal2_Res_Filled'].sum()) / (len(df) * 2) * 100)
resource_text = ' '.join(df['Goal1_Resources'].fillna('').astype(str)) + ' ' + ' '.join(df['Goal2_Resources'].fillna('').astype(str))
resource_words = jieba.lcut(resource_text)
resource_words = [w for w in resource_words if len(w) >= 2 and w not in stopwords and w != '无' and not re.match(r'^[a-zA-Z0-9]+$', w)]
resource_keyword_counts = Counter(resource_words).most_common(20)
resource_keyword_df = pd.DataFrame(resource_keyword_counts, columns=['Keyword', 'Frequency'])

completeness_dist = df['Completeness_Tier'].value_counts().reset_index()
completeness_dist.columns = ['Completeness_Tier', 'Count']

manager_supported_methods = ['Seek feedback from manager', 'Coaching/Mentoring', 'Networking & peer learning']
manager_support_count = sum(action_method_counts.get(m, 0) for m in manager_supported_methods)
total_action_methods = sum(action_method_counts.values())

# --- Goal cycle analysis ---
total_goal_dates = df['Total_Goal_Date'].dropna().dt.to_period('M').value_counts().sort_index().reset_index()
total_goal_dates.columns = ['Month', 'Count']
total_goal_dates['Month'] = total_goal_dates['Month'].astype(str)

mid_dates = df['Mid_Review_Date'].dropna().dt.to_period('M').value_counts().sort_index().reset_index()
mid_dates.columns = ['Month', 'Count']
mid_dates['Month'] = mid_dates['Month'].astype(str)

end_dates = df['Final_Review_Date'].dropna().dt.to_period('M').value_counts().sort_index().reset_index()
end_dates.columns = ['Month', 'Count']
end_dates['Month'] = end_dates['Month'].astype(str)

cycle_length_stats = pd.DataFrame({
    'Metric': ['Fill to Total Goal (days)', 'Mid to Final Review (days)', 'Final to Total Goal (days)'],
    'Mean': [
        round(df['Days_Fill_to_Total'].mean(), 1) if pd.notna(df['Days_Fill_to_Total'].mean()) else 0,
        round(df['Days_Mid_to_Final'].mean(), 1) if pd.notna(df['Days_Mid_to_Final'].mean()) else 0,
        round(df['Days_Final_to_Total'].mean(), 1) if pd.notna(df['Days_Final_to_Total'].mean()) else 0
    ],
    'Median': [
        round(df['Days_Fill_to_Total'].median(), 1) if pd.notna(df['Days_Fill_to_Total'].median()) else 0,
        round(df['Days_Mid_to_Final'].median(), 1) if pd.notna(df['Days_Mid_to_Final'].median()) else 0,
        round(df['Days_Final_to_Total'].median(), 1) if pd.notna(df['Days_Final_to_Total'].median()) else 0
    ],
    'Min': [
        int(df['Days_Fill_to_Total'].min()) if pd.notna(df['Days_Fill_to_Total'].min()) else 0,
        int(df['Days_Mid_to_Final'].min()) if pd.notna(df['Days_Mid_to_Final'].min()) else 0,
        int(df['Days_Final_to_Total'].min()) if pd.notna(df['Days_Final_to_Total'].min()) else 0
    ],
    'Max': [
        int(df['Days_Fill_to_Total'].max()) if pd.notna(df['Days_Fill_to_Total'].max()) else 0,
        int(df['Days_Mid_to_Final'].max()) if pd.notna(df['Days_Mid_to_Final'].max()) else 0,
        int(df['Days_Final_to_Total'].max()) if pd.notna(df['Days_Final_to_Total'].max()) else 0
    ]
})

anomalies = pd.DataFrame({
    'Anomaly_Type': [
        'Total goal date before fill date',
        'Mid review after final review',
        'Final review after total goal date'
    ],
    'Count': [
        int(df['Anomaly_Total_Before_Fill'].sum()),
        int(df['Anomaly_Mid_After_Final'].sum()),
        int(df['Anomaly_Final_After_Total'].sum())
    ]
})

# ============================================================
# Generate Charts
# ============================================================
print("Generating charts...")

def save_chart(fig, filename):
    path = os.path.join(CHART_DIR, filename)
    fig.savefig(path, dpi=150, bbox_inches='tight')
    plt.close(fig)
    return path

# Chart 1: Position distribution
fig, ax = plt.subplots(figsize=(8, 5))
ax.bar(position_dist['Position'], position_dist['Count'], color='steelblue')
ax.set_title('Position Distribution', fontsize=14, fontweight='bold')
ax.set_xlabel('Position')
ax.set_ylabel('Number of Employees')
plt.xticks(rotation=30, ha='right')
save_chart(fig, 'position_dist.png')

# Chart 2: Goal theme distribution
fig, ax = plt.subplots(figsize=(8, 5))
ax.barh(theme_dist['Theme'], theme_dist['Count'], color='coral')
ax.set_title('Development Goal Theme Distribution', fontsize=14, fontweight='bold')
ax.set_xlabel('Count')
save_chart(fig, 'theme_dist.png')

# Chart 3: Skill attribute distribution
fig, ax = plt.subplots(figsize=(6, 6))
ax.pie(skill_attr_dist['Count'], labels=skill_attr_dist['Skill_Type'], autopct='%1.1f%%', startangle=90)
ax.set_title('Skill Type Distribution (Soft vs Hard)', fontsize=14, fontweight='bold')
save_chart(fig, 'skill_attr_dist.png')

# Chart 4: Action coverage
fig, ax = plt.subplots(figsize=(8, 5))
x = np.arange(len(action_coverage))
width = 0.35
ax.bar(x - width/2, action_coverage['Goal1_Filled'], width, label='Goal 1', color='teal')
ax.bar(x + width/2, action_coverage['Goal2_Filled'], width, label='Goal 2', color='orange')
ax.set_xticks(x)
ax.set_xticklabels(action_coverage['Dimension'])
ax.set_ylabel('Number of Employees')
ax.set_title('70-20-10 Action Plan Coverage', fontsize=14, fontweight='bold')
ax.legend()
save_chart(fig, 'action_coverage.png')

# Chart 5: Completeness distribution
fig, ax = plt.subplots(figsize=(6, 6))
tier_order = [t for t in ['High', 'Medium', 'Low'] if t in completeness_dist['Completeness_Tier'].values]
cd_ordered = completeness_dist.set_index('Completeness_Tier').reindex(tier_order).reset_index()
colors_map = {'High': '#2ecc71', 'Medium': '#f39c12', 'Low': '#e74c3c'}
colors_list = [colors_map.get(t, 'gray') for t in cd_ordered['Completeness_Tier']]
ax.pie(cd_ordered['Count'], labels=cd_ordered['Completeness_Tier'], autopct='%1.1f%%', startangle=90, colors=colors_list)
ax.set_title('IDP Completeness Tier Distribution', fontsize=14, fontweight='bold')
save_chart(fig, 'completeness_dist.png')

# Chart 6: Total goal date timeline
fig, ax = plt.subplots(figsize=(10, 5))
ax.plot(total_goal_dates['Month'], total_goal_dates['Count'], marker='o', color='purple')
ax.set_title('Total Goal Achievement Date Distribution (by Month)', fontsize=14, fontweight='bold')
ax.set_xlabel('Month')
ax.set_ylabel('Number of Employees')
plt.xticks(rotation=45, ha='right')
save_chart(fig, 'total_goal_timeline.png')

# Chart 7: Region distribution (if available)
chart_region = None
if 'Region' in df.columns and num_regions > 0:
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(region_dist['Region'], region_dist['Count'], color='mediumpurple')
    ax.set_title('Region Distribution', fontsize=14, fontweight='bold')
    ax.set_xlabel('Region')
    ax.set_ylabel('Number of Employees')
    plt.xticks(rotation=30, ha='right')
    chart_region = save_chart(fig, 'region_dist.png')

# ============================================================
# Generate Excel Workbook
# ============================================================
print("Generating Excel workbook...")
excel_path = os.path.join(OUTPUT_DIR, 'IDP_Analysis_Results.xlsx')

# Prepare raw data sheet
raw_cols = ['Employee_ID', 'Employee_Name', 'Manager_Name', 'Position', 'Region', 'Store', 'Fill_Date',
            'Goal1_Type', 'Goal1_Description', 'Goal1_70_Plan', 'Goal1_70_Detail', 'Goal1_20_Plan', 'Goal1_20_Detail',
            'Goal1_10_Plan', 'Goal1_10_Detail', 'Goal1_Resources',
            'Goal2_Type', 'Goal2_Description', 'Goal2_70_Plan', 'Goal2_70_Detail', 'Goal2_20_Plan', 'Goal2_20_Detail',
            'Goal2_10_Plan', 'Goal2_10_Detail', 'Goal2_Resources',
            'Mobility_Accept', 'Mobility_Target', 'Total_Goal_Date', 'Mid_Review_Date', 'Final_Review_Date',
            'Goals_Total', 'Completeness_Score', 'Completeness_Pct', 'Completeness_Tier', 'Goal_Fill_Status']
# Filter to columns that exist
raw_cols = [c for c in raw_cols if c in df.columns]
df_raw = df[raw_cols].copy()

# Prepare employee score sheet
score_cols = ['Employee_ID', 'Employee_Name', 'Manager_Name', 'Position', 'Store', 'Goals_Total',
              'Completeness_Score', 'Completeness_Pct', 'Completeness_Tier',
              'Goal1_Filled', 'Goal1_70_Filled', 'Goal1_20_Filled', 'Goal1_10_Filled', 'Goal1_Res_Filled',
              'Goal2_Filled', 'Goal2_70_Filled', 'Goal2_20_Filled', 'Goal2_10_Filled', 'Goal2_Res_Filled',
              'Days_Fill_to_Total', 'Days_Mid_to_Final', 'Days_Final_to_Total']
score_cols = [c for c in score_cols if c in df.columns]
df_score = df[score_cols].copy()

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df_raw.to_excel(writer, sheet_name='Raw Data', index=False)
    position_dist.to_excel(writer, sheet_name='Basic Summary', startrow=0, startcol=0, index=False)
    manager_dist.to_excel(writer, sheet_name='Basic Summary', startrow=0, startcol=4, index=False)
    store_dist.to_excel(writer, sheet_name='Basic Summary', startrow=0, startcol=8, index=False)
    mobility_dist.to_excel(writer, sheet_name='Basic Summary', startrow=0, startcol=12, index=False)
    if 'Region' in df.columns:
        region_dist.to_excel(writer, sheet_name='Basic Summary', startrow=0, startcol=16, index=False)

    goal_completeness_dist.to_excel(writer, sheet_name='Goal Analysis', startrow=0, startcol=0, index=False)
    goal_type_combined.to_excel(writer, sheet_name='Goal Analysis', startrow=0, startcol=4, index=False)
    skill_attr_dist.to_excel(writer, sheet_name='Goal Analysis', startrow=0, startcol=8, index=False)
    theme_dist.to_excel(writer, sheet_name='Goal Analysis', startrow=0, startcol=12, index=False)
    keyword_df[['Keyword', 'Keyword_EN', 'Frequency']].to_excel(writer, sheet_name='Goal Analysis', startrow=0, startcol=16, index=False)

    action_coverage.to_excel(writer, sheet_name='Action Plan Analysis', startrow=0, startcol=0, index=False)
    action_method_df.to_excel(writer, sheet_name='Action Plan Analysis', startrow=0, startcol=7, index=False)
    resource_keyword_df.to_excel(writer, sheet_name='Action Plan Analysis', startrow=0, startcol=12, index=False)
    completeness_dist.to_excel(writer, sheet_name='Action Plan Analysis', startrow=0, startcol=18, index=False)

    total_goal_dates.to_excel(writer, sheet_name='Cycle Analysis', startrow=0, startcol=0, index=False)
    mid_dates.to_excel(writer, sheet_name='Cycle Analysis', startrow=0, startcol=4, index=False)
    end_dates.to_excel(writer, sheet_name='Cycle Analysis', startrow=0, startcol=8, index=False)
    cycle_length_stats.to_excel(writer, sheet_name='Cycle Analysis', startrow=0, startcol=12, index=False)
    anomalies.to_excel(writer, sheet_name='Cycle Analysis', startrow=0, startcol=19, index=False)

    df_score.to_excel(writer, sheet_name='Employee Detail Scores', index=False)

# Format Excel workbook
wb = openpyxl.load_workbook(excel_path)
header_fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
header_font = Font(bold=True)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

wb.save(excel_path)
print(f"Excel saved: {excel_path}")

# ============================================================
# Generate Markdown Report (English)
# ============================================================
print("Generating Markdown report...")
report_path = os.path.join(OUTPUT_DIR, 'IDP_Analysis_Report.md')

r = []
r.append("# Individual Development Plan (IDP) Analysis Report")
r.append("")
r.append(f"**Analysis Date:** {datetime.now().strftime('%Y-%m-%d')}")
r.append(f"**Data Source:** {os.path.basename(INPUT_FILE)}")
r.append(f"**Sample Size:** {num_employees} employees")
r.append("")
r.append("---")
r.append("")

# Executive Summary
r.append("## 1. Executive Summary")
r.append("")
r.append(f"- This analysis covers **{num_employees} employees** across **{num_stores} stores** in **{num_regions} regions**, managed by **{num_managers} managers**.")
position_summary = ', '.join([f"{row['Position']}: {row['Count']}" for _, row in position_dist.iterrows()])
r.append(f"- Position distribution: {position_summary}.")
total_goals_filled = int(df['Goals_Total'].sum())
r.append(f"- A total of **{total_goals_filled} development goals** were filled, averaging **{total_goals_filled / num_employees:.2f} goals per employee**.")
r.append(f"- Average IDP completeness score: **{df['Completeness_Score'].mean():.1f} / 12 ({df['Completeness_Pct'].mean():.1f}%)**.")
high_count = int((df['Completeness_Tier'] == 'High').sum())
med_count = int((df['Completeness_Tier'] == 'Medium').sum())
low_count = int((df['Completeness_Tier'] == 'Low').sum())
r.append(f"- Completeness breakdown: **{high_count} High**, **{med_count} Medium**, **{low_count} Low**.")
r.append("")

# Basic Situation
r.append("## 2. Basic Situation Analysis")
r.append("")
r.append("### 2.1 Workforce & Organization Overview")
r.append("")
r.append(f"| Metric | Value |")
r.append(f"|---|---|")
r.append(f"| Total Employees | {num_employees} |")
r.append(f"| Total Managers | {num_managers} |")
r.append(f"| Total Stores | {num_stores} |")
r.append(f"| Total Regions | {num_regions} |")
r.append(f"| Position Types | {num_positions} |")
r.append("")

r.append("### 2.2 Position Distribution")
r.append("")
r.append(position_dist.to_markdown(index=False))
r.append("")
r.append("![Position Distribution](charts/position_dist.png)")
r.append("")

if 'Region' in df.columns:
    r.append("### 2.3 Region Distribution")
    r.append("")
    r.append(region_dist.to_markdown(index=False))
    r.append("")
    r.append("![Region Distribution](charts/region_dist.png)")
    r.append("")

r.append("### 2.4 Manager Direct Reports (Top 10)")
r.append("")
r.append(manager_dist.head(10).to_markdown(index=False))
r.append("")

r.append("### 2.5 Store Distribution (Top 10)")
r.append("")
r.append(store_dist.head(10).to_markdown(index=False))
r.append("")

r.append("### 2.6 Mobility Willingness")
r.append("")
r.append(mobility_dist.to_markdown(index=False))
r.append("")

# IDP Goal Analysis
r.append("## 3. IDP Goal Analysis")
r.append("")
r.append("### 3.1 Goal Filling Completeness")
r.append("")
r.append(goal_completeness_dist.to_markdown(index=False))
r.append("")

r.append("### 3.2 Goal Type Distribution (Combined Goal 1 & Goal 2)")
r.append("")
r.append(goal_type_combined.to_markdown(index=False))
r.append("")

r.append("### 3.3 Skill Type Distribution (Soft vs Hard)")
r.append("")
r.append(skill_attr_dist.to_markdown(index=False))
r.append("")
r.append("![Skill Type Distribution](charts/skill_attr_dist.png)")
r.append("")

r.append("### 3.4 Goal Theme Distribution")
r.append("")
r.append(theme_dist.to_markdown(index=False))
r.append("")
r.append("![Goal Theme Distribution](charts/theme_dist.png)")
r.append("")

r.append("### 3.5 High-Frequency Keywords in Goal Descriptions (Top 20)")
r.append("")
r.append(keyword_df[['Keyword', 'Keyword_EN', 'Frequency']].head(20).to_markdown(index=False))
r.append("")

# Action Plan Analysis
r.append("## 4. Action Plan Analysis")
r.append("")
r.append("### 4.1 70-20-10 Learning Model Coverage")
r.append("")
r.append(action_coverage.to_markdown(index=False))
r.append("")
r.append("![Action Plan Coverage](charts/action_coverage.png)")
r.append("")

r.append("### 4.2 Common Action Methods Distribution")
r.append("")
r.append(action_method_df.to_markdown(index=False))
r.append("")

r.append("### 4.3 Resource Needs Analysis")
r.append("")
r.append(f"- Resource field fill rate: **{resource_filled_rate:.1f}%**")
r.append("- Top resource keywords:")
r.append("")
r.append(resource_keyword_df.to_markdown(index=False))
r.append("")

r.append("### 4.4 IDP Completeness Score Distribution")
r.append("")
r.append(completeness_dist.to_markdown(index=False))
r.append("")
r.append("![Completeness Distribution](charts/completeness_dist.png)")
r.append("")

r.append("### 4.5 Manager Support Analysis")
r.append("")
r.append(f"- Manager-dependent learning methods (coaching, seeking feedback, networking & peer learning) appeared **{manager_support_count} times**.")
if total_action_methods > 0:
    r.append(f"- This accounts for **{manager_support_count / total_action_methods * 100:.1f}%** of all action method records.")
r.append("")

# Goal Cycle Analysis
r.append("## 5. Goal Cycle Analysis")
r.append("")
r.append("### 5.1 Cycle Length Statistics")
r.append("")
r.append(cycle_length_stats.to_markdown(index=False))
r.append("")

r.append("### 5.2 Total Goal Achievement Date Distribution (by Month)")
r.append("")
r.append(total_goal_dates.to_markdown(index=False))
r.append("")
r.append("![Total Goal Timeline](charts/total_goal_timeline.png)")
r.append("")

r.append("### 5.3 Timeline Anomaly Check")
r.append("")
r.append(anomalies.to_markdown(index=False))
r.append("")

# Insights & Recommendations
r.append("## 6. Insights & Recommendations")
r.append("")
r.append("### 6.1 Workforce Insights")
r.append("")
manager_std = manager_dist['Direct_Reports'].std()
r.append(f"- Manager coverage is {'relatively balanced' if manager_std < 3 else 'uneven'}, with an average of {manager_dist['Direct_Reports'].mean():.1f} direct reports per manager (std: {manager_std:.1f}).")
yes_count = int((df['Mobility_Accept'] == 'Yes').sum())
no_count = int((df['Mobility_Accept'] == 'No').sum())
r.append(f"- Mobility willingness: **{yes_count} employees** willing to relocate, **{no_count} employees** declined.")
r.append("")

r.append("### 6.2 IDP Goal Insights")
r.append("")
theme_dist_valid = theme_dist[theme_dist['Theme'] != 'Not Filled']
if not theme_dist_valid.empty:
    top_theme = theme_dist_valid.iloc[0]['Theme']
    top_theme_count = int(theme_dist_valid.iloc[0]['Count'])
    r.append(f"- The most popular development theme is **{top_theme}** ({top_theme_count} occurrences). Consider designing targeted training or workshops around this theme.")
else:
    r.append("- No valid development goal theme data available.")

soft_count = int(skill_attr_dist[skill_attr_dist['Skill_Type'] == 'Soft Skill']['Count'].sum())
hard_count = int(skill_attr_dist[skill_attr_dist['Skill_Type'] == 'Hard Skill']['Count'].sum())
total_skill = soft_count + hard_count
if total_skill > 0:
    r.append(f"- Soft skill goals account for **{soft_count / total_skill * 100:.1f}%**, hard skill goals account for **{hard_count / total_skill * 100:.1f}%**.")

single_or_none = int(((df['Goals_Total'] == 0) | (df['Goals_Total'] == 1)).sum())
r.append(f"- **{single_or_none} employees** filled only one goal or no goals at all. Recommend pushing all employees to complete both goals.")
r.append("")

r.append("### 6.3 Action Plan Insights")
r.append("")
g1_70_rate = df['Goal1_70_Filled'].sum() / len(df) * 100
g2_70_rate = df['Goal2_70_Filled'].sum() / len(df) * 100
g1_10_rate = df['Goal1_10_Filled'].sum() / len(df) * 100
g2_10_rate = df['Goal2_10_Filled'].sum() / len(df) * 100
r.append(f"- 70% on-the-job practice has the highest fill rate (Goal 1: {g1_70_rate:.1f}%, Goal 2: {g2_70_rate:.1f}%), aligning with the IDP design principle.")
r.append(f"- 10% formal courses has the lowest fill rate (Goal 1: {g1_10_rate:.1f}%, Goal 2: {g2_10_rate:.1f}%). Recommend strengthening course recommendations and training resource alignment.")
r.append(f"- Average completeness score is **{df['Completeness_Score'].mean():.1f}/12**. Recommend targeted follow-up for employees scoring below 6.")
r.append("")

r.append("### 6.4 Goal Cycle Insights")
r.append("")
mean_fill_total = df['Days_Fill_to_Total'].mean()
mean_mid_final = df['Days_Mid_to_Final'].mean()
r.append(f"- Average goal cycle length is **{mean_fill_total:.0f} days**, with an average of **{mean_mid_final:.0f} days** from mid to final review.")
total_anomalies = int(anomalies['Count'].sum())
r.append(f"- **{total_anomalies} timeline anomalies** detected. Recommend reviewing with relevant managers to correct.")
r.append("")

r.append("### 6.5 Action Recommendations")
r.append("")
r.append("1. **Promote Dual-Goal Completion**: For employees who filled only one goal or none, managers should facilitate completing the second goal during 1-on-1 sessions.")
r.append("2. **Strengthen Course Resource Alignment**: The 10% formal course dimension has low fill rates. HR/L&D should curate a course catalog and provide targeted recommendations.")
r.append("3. **Establish Manager Support Mechanisms**: Encourage managers to proactively provide feedback, coaching, and networking resources to improve the quality of the 20% learning-from-others dimension.")
r.append("4. **Correct Timeline Anomalies**: Review records with illogical date sequences to ensure mid-review, final review, and total goal dates are logically consistent.")
r.append("5. **Track Completeness Quarterly**: Use IDP completeness as a quarterly management metric. Send timely reminders to employees with low completeness scores.")
r.append("")

# Appendix
r.append("## 7. Appendix")
r.append("")
r.append("### 7.1 Data Dictionary")
r.append("")
r.append("- **Completeness Score**: 12-point scale covering goal type, goal description, three action plan dimensions (70/20/10), and resource fields for both goals.")
r.append("- **70-20-10 Model**: 70% on-the-job practice, 20% learning from others, 10% formal courses.")
r.append("- **Soft Skills**: Communication, Planning & Organization, Result Orientation, etc.")
r.append("- **Hard Skills**: Sales Art, Customer Excellence, Brand & Culture, etc.")
r.append("")

r.append("### 7.2 Methodology")
r.append("")
r.append("- Data cleaning and statistics: Python pandas")
r.append("- Chinese keyword extraction: jieba")
r.append("- Charts: matplotlib")
r.append("- Excel generation: openpyxl")
r.append("")

with open(report_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(r))

print(f"Report saved: {report_path}")
print("All outputs generated successfully.")
